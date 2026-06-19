"""Exporta la conciliación a Excel para finanzas."""
from __future__ import annotations

from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from . import config

MONEY = {"costo", "sale_price", "sobrepeso_cobrado", "factura_cliente", "ingreso", "margen",
         "Costo", "Ingreso", "Margen", "Costo_DHL", "Falta_cobrar"}


def _sheet(ws, headers, rows):
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    for ci, h in enumerate(headers, 1):
        L = get_column_letter(ci)
        if h in MONEY:
            for cell in ws[L][1:]:
                cell.number_format = "#,##0.00"
        ws.column_dimensions[L].width = max(12, min(36, len(str(h)) + 2))
    ws.freeze_panes = "A2"


def export(con, out_path: str | None = None) -> str:
    config.EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    out = out_path or str(config.EXPORTS_DIR / "conciliacion.xlsx")
    wb = Workbook()

    # Resumen por estatus
    ws = wb.active
    ws.title = "Resumen_estatus"
    _sheet(ws, ["estatus", "guias", "costo", "ingreso", "margen"], con.execute("""
        SELECT estatus, count(*), round(sum(costo),2), round(sum(ingreso),2),
               round(sum(ingreso)-sum(costo),2)
        FROM reconciliacion GROUP BY estatus ORDER BY 3 DESC""").fetchall())

    # Por cliente real
    _sheet(wb.create_sheet("Por_cliente"),
           ["cliente_real", "tipo", "guias", "Costo", "Ingreso", "Margen", "margen_pct"],
           con.execute("""
        SELECT cliente_real, any_value(tipo), count(*),
               round(sum(costo),2) c, round(sum(ingreso),2) i,
               round(sum(ingreso)-sum(costo),2),
               round(100*(sum(ingreso)-sum(costo))/nullif(sum(costo),0),1)
        FROM reconciliacion GROUP BY cliente_real ORDER BY c DESC""").fetchall())

    # Por mes de envío (devengado)
    _sheet(wb.create_sheet("Por_mes_devengado"),
           ["mes_envio", "guias", "Costo", "Ingreso", "Margen"],
           con.execute("""
        SELECT mes_envio, count(*), round(sum(costo),2), round(sum(ingreso),2),
               round(sum(ingreso)-sum(costo),2)
        FROM reconciliacion GROUP BY mes_envio ORDER BY mes_envio""").fetchall())

    # Falta cobrar / problemas (lo accionable)
    _sheet(wb.create_sheet("Por_cobrar"),
           ["cliente_real", "estatus", "guias", "Costo_DHL", "Falta_cobrar"],
           con.execute("""
        SELECT cliente_real, estatus, count(*), round(sum(costo),2),
               round(sum(costo) - sum(coalesce(ingreso,0)),2) falta
        FROM reconciliacion
        WHERE estatus IN ('Falta cobrar (credito)','Sobrepeso pendiente','Cobrado bajo costo')
        GROUP BY cliente_real, estatus ORDER BY falta DESC""").fetchall())

    wb.save(out)
    return out
