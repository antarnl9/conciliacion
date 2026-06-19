"""Parser del formato Acre de DHL Express México.

Gotchas reales (de la auditoría):
- 3 variantes de esquema entre archivos -> SIEMPRE mapear por encabezado, no por posición.
- Columnas de recargo varían: FF, NX, OO, YB, YE, YK, YY. Se suman todas las presentes.
- Enums/valores en español. Importe Neto incluye IVA.
- Retorno = Referencia tipo `^RT\\d{10}$` (los 10 dígitos = guía de venta original).
"""
from __future__ import annotations

import re
from datetime import datetime
from typing import Iterator

from openpyxl import load_workbook

from .base import LineaFactura

RECARGO_COLS = {"FF", "NX", "OO", "YB", "YE", "YK", "YY"}
RT_RE = re.compile(r"RT\d{10}")


def _num(v) -> float:
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def _to_date(v):
    return v.date() if isinstance(v, datetime) else None


def _txt(v):
    if v is None:
        return None
    s = re.sub(r"\s+", " ", str(v).strip())
    return s or None


def parse_dhl(path: str, sheet: str | None = None) -> Iterator[LineaFactura]:
    """Genera LineaFactura desde un archivo Acre de DHL (streaming, baja memoria)."""
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else None for h in next(it)]
    idx = {h: i for i, h in enumerate(header) if h}

    def col(name):
        return idx.get(name)

    i_guia = col("No.De Guia")
    i_net = col("Importe Neto")
    i_cta = col("No.De Cuenta")
    i_ref = col("Referencia")
    i_prod = col("Producto")
    i_org = col("Org")
    i_des = col("Des")
    i_pza = col("Pza")
    i_kg = col("Kilos")
    i_fenv = col("Fecha Envio")
    i_ffac = col("Fecha Factura")
    i_fac = col("No. Factura")
    i_flete = col("Flete")
    i_seg = col("Seguro")
    i_desc = col("Descuento")
    i_iva = col("Imp.I.V.A.")
    i_mon = col("Moneda")
    i_rem = col("Remitente")
    i_dst = col("Destinatario")
    i_zona = col("Zona")
    rec_idx = [idx[c] for c in idx if c in RECARGO_COLS]

    archivo = path.rsplit("/", 1)[-1]
    for row in it:
        if row is None or all(v is None for v in row):
            continue

        def g(i):
            return row[i] if i is not None and i < len(row) else None

        guia = _txt(g(i_guia))
        if not guia:
            continue
        ref = _txt(g(i_ref))
        es_ret = bool(ref and RT_RE.fullmatch(ref.upper()))
        recargos = sum(_num(row[i]) for i in rec_idx if i < len(row))
        piezas = g(i_pza)
        yield LineaFactura(
            carrier="dhl",
            guia=guia,
            importe_neto=_num(g(i_net)),
            archivo_origen=archivo,
            cuenta=_txt(g(i_cta)),
            referencia=ref,
            producto=_txt(g(i_prod)),
            origen=_txt(g(i_org)),
            destino=_txt(g(i_des)),
            piezas=int(piezas) if isinstance(piezas, (int, float)) else None,
            kilos=_num(g(i_kg)),
            fecha_envio=_to_date(g(i_fenv)),
            fecha_factura=_to_date(g(i_ffac)),
            no_factura=_txt(g(i_fac)),
            flete=_num(g(i_flete)),
            seguro=_num(g(i_seg)),
            descuento=_num(g(i_desc)),
            recargos=recargos,
            iva=_num(g(i_iva)),
            moneda=_txt(g(i_mon)) or "MXN",
            remitente=_txt(g(i_rem)),
            destinatario=_txt(g(i_dst)),
            es_retorno=es_ret,
            zona=_txt(g(i_zona)),
        )
    wb.close()
