"""Lector de facturas reales a cliente (crédito): columnas guia + total [+ cliente]."""
from __future__ import annotations

from typing import Iterator

from openpyxl import load_workbook


def iter_facturas_cliente(path: str) -> Iterator[tuple]:
    """Genera (guia, cliente, total, archivo_origen) de cada hoja con guia+total."""
    archivo = path.rsplit("/", 1)[-1]
    wb = load_workbook(path, read_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        it = ws.iter_rows(values_only=True)
        try:
            hdr = [str(h).strip() if h is not None else None for h in next(it)]
        except StopIteration:
            continue
        idx = {h: i for i, h in enumerate(hdr) if h}
        ig = idx.get("No.De Guia", idx.get("guia"))
        itot = idx.get("Total neto", idx.get("Total"))
        icli = idx.get("Remitente", idx.get("Cliente"))
        if ig is None or itot is None:
            continue
        for row in it:
            if row is None:
                continue
            g = row[ig] if ig < len(row) else None
            if g is None:
                continue
            t = row[itot] if itot is not None and itot < len(row) else None
            c = row[icli] if icli is not None and icli < len(row) else None
            yield (str(g).strip(),
                   str(c).strip() if c else None,
                   float(t) if isinstance(t, (int, float)) else 0.0,
                   archivo)
    wb.close()
