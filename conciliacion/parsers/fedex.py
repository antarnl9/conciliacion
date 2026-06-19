"""Parser del archivo de facturación de FedEx (export 'terradata', 1 hoja, ~63 cols).

- Mapeo por encabezado. Costo = `Total_Facturado` (neto, ya con IVA).
- Peso = `Peso_facturado` (kg). Zona = `Rate_Zona` (ej. 'DL 7').
- Las fechas vienen como datetime de Excel.
"""
from __future__ import annotations

import re
from datetime import datetime
from typing import Iterator

from openpyxl import load_workbook

from .base import LineaFactura

RT_RE = re.compile(r"RT\d{10}")


def _num(v) -> float:
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def _to_date(v):
    return v.date() if isinstance(v, datetime) else None


def _txt(v):
    if v is None:
        return None
    s = re.sub(r"\s+", " ", str(v).strip())
    return s or None


_MAP = {
    "guia": "Guia", "net": "Total_Facturado", "cta": "Cuenta", "ref": "Referencia",
    "prod": "Servicio", "org": "Ciudad_Shipper", "des": "Ciudad_Destino", "pza": "Volumen",
    "kg": "Peso_facturado", "fenv": "Fecha_Envio", "ffac": "Fecha_Factura", "fac": "Factura",
    "flete": "Flete", "desc": "Descuento", "fuel": "Fuel_Surcharge", "iva": "MEXICO_IVA_FRT",
    "mon": "Moneda", "rem": "Nombre_quien_Envia", "dst": "Nombre_Destinatario", "zona": "Rate_Zona",
}


def parse_fedex(path: str, sheet: str | None = None) -> Iterator[LineaFactura]:
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else None for h in next(it)]
    idx = {h: i for i, h in enumerate(header) if h}
    c = {k: idx.get(v) for k, v in _MAP.items()}
    archivo = path.rsplit("/", 1)[-1]

    for row in it:
        if row is None or all(v is None for v in row):
            continue

        def g(i):
            return row[i] if i is not None and i < len(row) else None

        guia = _txt(g(c["guia"]))
        if not guia:
            continue
        ref = _txt(g(c["ref"]))
        pza = g(c["pza"])
        yield LineaFactura(
            carrier="fedex", guia=guia, importe_neto=_num(g(c["net"])), archivo_origen=archivo,
            cuenta=_txt(g(c["cta"])), referencia=ref, producto=_txt(g(c["prod"])),
            origen=_txt(g(c["org"])), destino=_txt(g(c["des"])),
            piezas=int(pza) if isinstance(pza, (int, float)) else None, kilos=_num(g(c["kg"])),
            fecha_envio=_to_date(g(c["fenv"])), fecha_factura=_to_date(g(c["ffac"])),
            no_factura=_txt(g(c["fac"])), flete=_num(g(c["flete"])), descuento=_num(g(c["desc"])),
            recargos=_num(g(c["fuel"])), iva=_num(g(c["iva"])), moneda=_txt(g(c["mon"])) or "MXN",
            remitente=_txt(g(c["rem"])), destinatario=_txt(g(c["dst"])),
            es_retorno=bool(ref and RT_RE.fullmatch(ref.upper())), zona=_txt(g(c["zona"])),
        )
    wb.close()
