"""Parser del Acre de Paquete Express (1 hoja, ~80 cols).

- Mapeo por encabezado. Costo = `Total` (neto, ya con IVA). Subtotal = Flete + RAD + Otros.
- Peso = `Peso` (kg). Zona = `Tarifa` (código de tarifa, ej. 'T0').
- Las fechas vienen como texto `DD/MM/YYYY HH:MM[:SS]`.
"""
from __future__ import annotations

import re
from datetime import datetime
from typing import Iterator

from openpyxl import load_workbook

from .base import LineaFactura


def _num(v) -> float:
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def _to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if v is None:
        return None
    s = str(v).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _txt(v):
    if v is None:
        return None
    s = re.sub(r"\s+", " ", str(v).strip())
    return s or None


_MAP = {
    # CRUCE: el shipment_number en ClickHouse es el "Rastreo", NO la "Guía" interna de Paquete Express.
    "guia": "Rastreo", "cta": "Guía", "net": "Total", "ref": "Referencia", "prod": "Tipo Servicio",
    "org": "Plaza orig.", "des": "Ciudad dest.", "pza": "Cantidad", "kg": "Peso",
    "fenv": "Fecha envío", "ffac": "Fecha factura", "fac": "Factura", "flete": "Flete",
    "seg": "Seguro", "iva": "IVA", "rem": "Cliente origen", "dst": "Cliente destino",
    "zona": "Tarifa", "otros": "Otros", "rad": "RAD",
}


def parse_paquete_express(path: str, sheet: str | None = None,
                          carrier: str = "paquete_express") -> Iterator[LineaFactura]:
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else None for h in next(it)]
    idx = {h: i for i, h in enumerate(header) if h}
    c = {k: idx.get(v) for k, v in _MAP.items()}
    archivo = path.rsplit("/", 1)[-1]

    for row in it:
        if row is None or all(v is None for v in row):
            continue

        def g(i):
            return row[i] if i is not None and i < len(row) else None

        guia = _txt(g(c["guia"]))
        if not guia:
            continue
        pza = g(c["pza"])
        recargos = _num(g(c["otros"])) + _num(g(c["rad"]))
        yield LineaFactura(
            carrier=carrier, guia=guia, importe_neto=_num(g(c["net"])), archivo_origen=archivo,
            cuenta=_txt(g(c["cta"])), referencia=_txt(g(c["ref"])), producto=_txt(g(c["prod"])),
            origen=_txt(g(c["org"])), destino=_txt(g(c["des"])),
            piezas=int(pza) if isinstance(pza, (int, float)) else None, kilos=_num(g(c["kg"])),
            fecha_envio=_to_date(g(c["fenv"])), fecha_factura=_to_date(g(c["ffac"])),
            no_factura=_txt(g(c["fac"])), flete=_num(g(c["flete"])), seguro=_num(g(c["seg"])),
            recargos=recargos, iva=_num(g(c["iva"])), moneda="MXN",
            remitente=_txt(g(c["rem"])), destinatario=_txt(g(c["dst"])),
            es_retorno=False, zona=_txt(g(c["zona"])),
        )
    wb.close()
