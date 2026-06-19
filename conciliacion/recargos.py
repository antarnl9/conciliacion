"""Captura el costo de cada recargo por guía desde el Acre, según `recargos_mapeo`.

El usuario liga cada concepto (ej. 'Zona extendida') a una columna del Acre (ej. 'ODA' en FedEx).
Esta función hace una segunda pasada sobre el Excel y guarda (carrier, guia, concepto, monto)
en `factura_recargos` para las guías que traen ese recargo (> 0).
"""
from __future__ import annotations

from openpyxl import load_workbook

# Columna de la guía por paquetería (misma llave que usa el parser para el cruce).
_GUIA_COL = {"dhl": "No.De Guia", "fedex": "Guia",
             "paquete_express": "Rastreo", "paquete_express_2": "Rastreo"}


def _num(v) -> float:
    try:
        return float(v) if v not in (None, "") else 0.0
    except (TypeError, ValueError):
        return 0.0


def ingest_recargos(con, path: str, carrier: str) -> int:
    """Lee los recargos mapeados del Acre y los inserta en factura_recargos. Devuelve # filas."""
    mapeo = con.execute("SELECT concepto, columna FROM recargos_mapeo WHERE carrier=?", [carrier]).fetchall()
    con.execute("DELETE FROM factura_recargos WHERE carrier=?", [carrier])
    if not mapeo:
        return 0

    wb = load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else None for h in next(it)]
    idx = {h: i for i, h in enumerate(header) if h}
    gi = idx.get(_GUIA_COL.get(carrier, "No.De Guia"))
    cols = [(concepto, idx.get(columna)) for concepto, columna in mapeo]

    buf, total = [], 0
    for row in it:
        if gi is None or gi >= len(row):
            continue
        guia = row[gi]
        if guia is None:
            continue
        guia = str(guia).strip()
        if not guia:
            continue
        for concepto, ci in cols:
            if ci is None or ci >= len(row):
                continue
            monto = _num(row[ci])
            if monto:                      # solo guías que sí traen el recargo
                buf.append((carrier, guia, concepto, monto))
        if len(buf) >= 50_000:
            con.executemany("INSERT INTO factura_recargos VALUES (?,?,?,?)", buf)
            total += len(buf); buf.clear()
    if buf:
        con.executemany("INSERT INTO factura_recargos VALUES (?,?,?,?)", buf)
        total += len(buf)
    wb.close()
    return total
