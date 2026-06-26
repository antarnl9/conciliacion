"""Genera el xlsx de tarifa de un cliente de crédito con el mismo formato
visual que los archivos de las paqueterías:

- Cliente + paquetería + servicio en el encabezado.
- Matriz peso (filas) × zona (columnas) con el precio FINAL al cliente
  (costo × (1+fuel) × (1+margen) × (1+IVA)).
- 1 sub-tabla por servicio (DHL: "ECONOMY SELECT DOMESTIC" y "EXPRESS DOMESTIC").
- Información de vigencia + margen + combustible + IVA aplicado.

El xlsx se sube a Supabase Storage en
`tarifas/clientes/<seller_id>/<YYYY-MM-DD>_v<N>.xlsx` y queda registrado en
DuckDB en `tarifas_cliente_archivos` para auditoría histórica.
"""
from __future__ import annotations

import io
import datetime as dt
from collections import defaultdict
from typing import Iterable

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import config


# ---- nombres "humanos" por (carrier, codigo de servicio) -------------------
SERVICIO_NOMBRE = {
    ("dhl", "G"): "ECONOMY SELECT DOMESTIC",
    ("dhl", "N"): "EXPRESS DOMESTIC",
    ("fedex", "Express Saver"): "EXPRESS SAVER",
    ("paquete_express", "Standard"): "STANDARD",
}


def servicio_nombre(carrier: str, codigo: str) -> str:
    return SERVICIO_NOMBRE.get((carrier, codigo)) or codigo or "GENERAL"


# ---- estilos compartidos ---------------------------------------------------
_HEAD_FILL = PatternFill("solid", fgColor="1F2937")
_HEAD_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F2937")
_SUBTITLE_FONT = Font(name="Calibri", size=11, bold=True, color="DB3B2B")
_TXT_FONT = Font(name="Calibri", size=10, color="1F2937")
_PRICE_FONT = Font(name="Calibri", size=10, color="1F2937")
_GRID = Border(
    left=Side(border_style="thin", color="E5E7EB"),
    right=Side(border_style="thin", color="E5E7EB"),
    top=Side(border_style="thin", color="E5E7EB"),
    bottom=Side(border_style="thin", color="E5E7EB"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def _set(ws, row: int, col: int, value, *, font=None, fill=None, border=None, align=None, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if font: c.font = font
    if fill: c.fill = fill
    if border: c.border = border
    if align: c.alignment = align
    if fmt: c.number_format = fmt
    return c


# ---- construcción de matriz por servicio -----------------------------------
def _build_matriz(rows: Iterable[dict], iva: float) -> tuple[list, list[tuple[float, float]]]:
    """De la salida de tarifa_preview, arma (zonas, pesos) y un dict precio[(zona,pesoKey)]=precio.

    rows: lista de {servicio, zona, peso_min, peso_max, costo, fuel, precio}
    """
    zonas = sorted(set(str(r["zona"]) for r in rows), key=lambda z: (not z.isdigit(), int(z) if z.isdigit() else 0, z))
    pesos = []
    seen = set()
    for r in sorted(rows, key=lambda r: (r.get("peso_min") or 0)):
        key = (r.get("peso_min"), r.get("peso_max"))
        if key not in seen:
            seen.add(key)
            pesos.append(key)
    return zonas, pesos


def _precio_label(price: float | None) -> float | str:
    if price is None:
        return ""
    return round(float(price), 2)


def _peso_label(peso_min, peso_max) -> str:
    if peso_min is None and peso_max is None:
        return "—"
    a = peso_min
    b = peso_max
    if a is not None and b is not None and (b - a) == 1 and float(a).is_integer():
        return f"{int(a)} kg"
    return f"{a if a is not None else 0} a {b if b is not None else '∞'}"


def build_xlsx(
    *,
    seller_id: int,
    cliente: str,
    carrier: str,
    metodo: str,
    margen: float | None,
    iva: float,
    vigencia_desde: str | None,
    vigencia_hasta: str | None,
    rows: list[dict],
    fuel_por_servicio: dict[str, float] | None = None,
) -> bytes:
    """Genera el xlsx en memoria y devuelve los bytes."""
    wb = openpyxl.Workbook()
    # Quitar la hoja default que crea openpyxl
    if wb.active and wb.active.title == "Sheet":
        wb.remove(wb.active)

    # Agrupar filas por servicio
    by_svc: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        by_svc[str(r.get("servicio") or "GENERAL")].append(r)

    fuel_por_servicio = fuel_por_servicio or {}

    if not by_svc:
        # Si no hay filas, genera una hoja vacía explicando el porqué
        ws = wb.create_sheet(title="Sin tarifa")
        _set(ws, 1, 1, f"Cliente: {cliente or seller_id}", font=_TITLE_FONT)
        _set(ws, 3, 1, f"Sin tarifa configurada para {carrier.upper()} aún.", font=_TXT_FONT)
        _set(ws, 5, 1, "Activa al cliente, configura método y margen, sube tarifa de paquetería.", font=_TXT_FONT)
        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue()

    # Una hoja por servicio
    for svc_code, svc_rows in by_svc.items():
        nombre_humano = servicio_nombre(carrier, svc_code)
        sheet_name = f"{carrier.upper()} {svc_code}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        # ----- encabezado -----
        _set(ws, 1, 1, f"{carrier.upper()} México", font=_TITLE_FONT)
        _set(ws, 2, 1, f"Cliente: {cliente or seller_id}", font=_SUBTITLE_FONT)
        _set(ws, 3, 1, nombre_humano, font=_SUBTITLE_FONT)

        vd = vigencia_desde or "—"
        vh = vigencia_hasta or "—"
        _set(ws, 4, 1, f"Vigencia: {vd} → {vh}", font=_TXT_FONT)
        margen_str = f"{(margen or 0) * 100:.2f}%" if margen is not None else "—"
        fuel_pct = fuel_por_servicio.get(svc_code) or 0.0
        _set(
            ws, 5, 1,
            f"Método: {metodo}  ·  Margen aplicado: {margen_str}  ·  Combustible: {fuel_pct*100:.2f}%  ·  IVA: {iva*100:.0f}%",
            font=_TXT_FONT,
        )
        _set(ws, 6, 1, "Precio final incluye combustible vigente, margen del cliente e IVA.", font=Font(italic=True, size=9, color="6B7280"))

        # ----- matriz -----
        head_row = 8
        zonas, pesos = _build_matriz(svc_rows, iva)
        # Header columna 1: "Kg"
        _set(ws, head_row, 1, "Kg", font=_HEAD_FONT, fill=_HEAD_FILL, border=_GRID, align=_CENTER)
        for i, z in enumerate(zonas):
            _set(ws, head_row, 2 + i, f"Zona {z}", font=_HEAD_FONT, fill=_HEAD_FILL, border=_GRID, align=_CENTER)

        # Index precios por (zona, peso_min, peso_max)
        idx = {(str(r["zona"]), r.get("peso_min"), r.get("peso_max")): r.get("precio") for r in svc_rows}

        for i, peso_key in enumerate(pesos):
            row_num = head_row + 1 + i
            _set(ws, row_num, 1, _peso_label(peso_key[0], peso_key[1]),
                 font=Font(bold=True, size=10, color="1F2937"), border=_GRID, align=_LEFT,
                 fill=PatternFill("solid", fgColor="F4F8FF" if i % 2 == 0 else "FFFFFF"))
            for j, z in enumerate(zonas):
                price = idx.get((z, peso_key[0], peso_key[1]))
                _set(
                    ws, row_num, 2 + j, _precio_label(price),
                    font=_PRICE_FONT, border=_GRID, align=_RIGHT,
                    fmt='"$"#,##0.00',
                    fill=PatternFill("solid", fgColor="F4F8FF" if i % 2 == 0 else "FFFFFF"),
                )

        # Ancho columnas
        ws.column_dimensions["A"].width = 18
        for i in range(len(zonas)):
            ws.column_dimensions[get_column_letter(2 + i)].width = 12

        # Freeze panes en encabezado + columna Kg
        ws.freeze_panes = ws.cell(row=head_row + 1, column=2)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def storage_path(seller_id: int, version: int, when: dt.date | None = None) -> str:
    when = when or dt.date.today()
    return f"tarifas/clientes/{seller_id}/{when.isoformat()}_v{version}.xlsx"
