"""API FastAPI del motor de conciliación + UI web para finanzas.

Levantar:  .venv/bin/uvicorn conciliacion.api.main:app --port 8770
Abrir:     http://localhost:8770
"""
from __future__ import annotations

import datetime as dt
import shutil
from pathlib import Path

from fastapi import FastAPI, UploadFile, File, Query, Request
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles

from .. import config, db, ch_sync, reconcile, export, supa, pricing
from ..parsers import PARSERS
from . import jobs

UI_DIR = config.ROOT / "ui"            # .../conciliacion/ui (local) y /app/ui (Docker)
UPLOADS = config.DATA_DIR / "uploads"

app = FastAPI(title="Conciliación T1", version="0.1.0")


@app.middleware("http")
async def auth_mw(request: Request, call_next):
    """Protege /api/* con el login de Supabase (si está configurado)."""
    path = request.url.path
    if not path.startswith("/api/") or path in ("/api/config", "/api/health") or not supa.enabled():
        return await call_next(request)
    auth = request.headers.get("authorization", "")
    token = auth[7:] if auth.lower().startswith("bearer ") else (request.query_params.get("token") or "")
    if not supa.verify_token(token):
        return JSONResponse({"error": "no autorizado"}, status_code=401)
    return await call_next(request)


@app.get("/api/config")
def app_config():
    """Config pública para el frontend (URL + llave pública de Supabase)."""
    s = config.supabase_settings()
    return {"auth_enabled": s["enabled"], "supabase_url": s["url"], "supabase_key": s["public"],
            "iva": config.IVA_DEFAULT}


# ---------- helpers ----------
def _save_upload(up: UploadFile, subdir: str) -> Path:
    dest_dir = UPLOADS / subdir
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / up.filename
    with dest.open("wb") as f:
        shutil.copyfileobj(up.file, f)
    return dest


def _rows(con, sql, params=None):
    cur = con.execute(sql, params or [])
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, r)) for r in cur.fetchall()]


def _fecha(desde: str, hasta: str, col: str = "mes_envio"):
    """Devuelve (clausulas, params) para filtrar por rango de mes_envio (YYYY-MM)."""
    cl, p = [], []
    if desde:
        cl.append(f"{col} >= ?"); p.append(desde)
    if hasta:
        cl.append(f"{col} <= ?"); p.append(hasta)
    return cl, p


# ---------- API ----------
@app.get("/api/status")
def status():
    con = db.connect()
    try:
        tablas = {t: con.execute(f"SELECT count(*) FROM {t}").fetchone()[0]
                  for t in ("facturas_carrier", "facturas_cliente", "ch_shipments", "ch_sobrepeso")}
        reconciled = con.execute("SELECT count(*) FROM reconciliacion").fetchone()[0]
        kpis, resumen = None, []
        if reconciled:
            resumen = _rows(con, """
                SELECT estatus, count(*) guias, round(sum(costo),2) costo,
                       round(sum(ingreso),2) ingreso, round(sum(ingreso)-sum(costo),2) margen
                FROM reconciliacion GROUP BY estatus ORDER BY costo DESC""")
            k = con.execute("""
                SELECT round(sum(costo),2) costo,
                       -- ingreso/utilidad REALIZADOS = solo guías que ya tienen costo de paquetería
                       round(sum(CASE WHEN costo IS NOT NULL THEN ingreso ELSE 0 END),2) ingreso,
                       round(sum(CASE WHEN costo IS NOT NULL AND ingreso IS NOT NULL THEN ingreso - costo ELSE 0 END),2) margen,
                       round(sum(CASE WHEN estatus IN ('Falta cobrar (credito)','Sobrepeso pendiente')
                              THEN costo - coalesce(ingreso,0) ELSE 0 END),2) por_cobrar,
                       -- cobrado pero la paquetería aún no factura el costo (devengado)
                       round(sum(CASE WHEN costo IS NULL THEN coalesce(ingreso,0) ELSE 0 END),2) por_devengar
                FROM reconciliacion""").fetchone()
            kpis = {"costo": k[0], "ingreso": k[1], "margen": k[2], "por_cobrar": k[3],
                    "por_devengar": k[4], "guias": reconciled}
        return {"tablas": tablas, "reconciled": reconciled, "kpis": kpis,
                "resumen": resumen, "jobs": jobs.snapshot()}
    finally:
        con.close()


@app.post("/api/upload/carrier")
def upload_carrier(carrier: str = Query("dhl"), reset: bool = Query(False),
                   file: UploadFile = File(...)):
    parser = PARSERS.get(carrier)
    if parser is None:
        return JSONResponse({"error": f"carrier no soportado: {carrier}"}, 400)
    path = _save_upload(file, "carrier")

    def _job():
        con = db.connect()
        try:
            if reset:
                con.execute("DELETE FROM facturas_carrier WHERE carrier = ?", [carrier])
            n = db.insert_facturas_carrier(con, parser(str(path)))
            return f"{n:,} lineas cargadas de {file.filename}"
        finally:
            con.close()

    return jobs.run(f"ingest-carrier:{file.filename}", _job)


@app.post("/api/upload/cliente")
def upload_cliente(reset: bool = Query(False), file: UploadFile = File(...)):
    path = _save_upload(file, "cliente")

    def _job():
        from ..parsers.cliente import iter_facturas_cliente
        con = db.connect()
        try:
            if reset:
                con.execute("DELETE FROM facturas_cliente")
            total = db.insert_facturas_cliente(con, iter_facturas_cliente(str(path)))
            return f"{total:,} facturas de cliente cargadas"
        finally:
            con.close()

    return jobs.run(f"ingest-cliente:{file.filename}", _job)


@app.post("/api/sync-ch")
def sync_ch(desde: str = Query("2025-09-01"), carrier: str = Query("dhl")):
    def _job():
        con = db.connect()
        try:
            n1 = ch_sync.sync_shipments(con, carrier=carrier, desde=desde)
            n2 = ch_sync.sync_sobrepeso(con, desde=desde)
            n3 = ch_sync.sync_recargas(con, desde=desde)
            return f"{n1:,} guias + {n2:,} sobrepeso + {n3:,} clientes con recargas"
        finally:
            con.close()

    return jobs.run("sync-ch", _job)


@app.post("/api/reconcile")
def do_reconcile():
    def _job():
        con = db.connect()
        try:
            n = reconcile.build(con)
            return f"{n:,} guias conciliadas"
        finally:
            con.close()

    return jobs.run("reconcile", _job)


@app.get("/api/clientes/prepago")
def clientes_prepago(limit: int = 500, desde: str = "", hasta: str = ""):
    fc, fp = _fecha(desde, hasta, "r.mes_envio")
    where = "WHERE r.tipo = 'prepago'" + "".join(" AND " + c for c in fc)
    con = db.connect()
    try:
        return _rows(con, f"""
            SELECT r.cliente_real,
                   count(*) guias,
                   round(sum(r.costo),2) costo,
                   round(sum(r.ingreso),2) ingreso,
                   round(sum(r.ingreso)-sum(r.costo),2) margen,
                   round(100*(sum(r.ingreso)-sum(r.costo))/nullif(sum(r.costo),0),1) margen_pct,
                   round(sum(r.sobrepeso_cobrado),2) sobrepeso,
                   round(sum(CASE WHEN r.es_retorno THEN r.costo ELSE 0 END),2) retornos,
                   count(*) FILTER (WHERE r.es_retorno) guias_retorno,
                   round(any_value(cr.recargas),2) recargas
            FROM reconciliacion r
            LEFT JOIN ch_recargas cr ON cr.seller_id = r.seller_id
            {where}
            GROUP BY r.cliente_real ORDER BY costo DESC LIMIT ?""", fp + [limit])
    finally:
        con.close()


@app.get("/api/clientes/credito")
def clientes_credito(limit: int = 500, desde: str = "", hasta: str = ""):
    fc, fp = _fecha(desde, hasta)
    where = "WHERE tipo = 'credito'" + "".join(" AND " + c for c in fc)
    con = db.connect()
    try:
        return _rows(con, f"""
            SELECT cliente_real,
                   count(*) guias,
                   round(sum(costo),2) costo,
                   round(sum(coalesce(ingreso,0)),2) facturado,
                   round(sum(CASE WHEN ingreso IS NULL THEN costo ELSE 0 END),2) falta_cobrar,
                   round(sum(sobrepeso_cobrado),2) sobrepeso,
                   round(sum(CASE WHEN es_retorno THEN costo ELSE 0 END),2) retornos,
                   count(*) FILTER (WHERE es_retorno) guias_retorno
            FROM reconciliacion {where}
            GROUP BY cliente_real ORDER BY costo DESC LIMIT ?""", fp + [limit])
    finally:
        con.close()


@app.get("/api/guias")
def guias(limit: int = 100, offset: int = 0, tipo: str = "", estatus: str = "",
          cliente: str = "", q: str = "", desde: str = "", hasta: str = ""):
    where, params = [], []
    if tipo:
        where.append("tipo = ?"); params.append(tipo)
    if estatus:
        where.append("estatus = ?"); params.append(estatus)
    if cliente:
        where.append("cliente_real ILIKE ?"); params.append(f"%{cliente}%")
    if q:
        where.append("guia ILIKE ?"); params.append(f"%{q}%")
    if desde:
        where.append("mes_envio >= ?"); params.append(desde)
    if hasta:
        where.append("mes_envio <= ?"); params.append(hasta)
    wsql = ("WHERE " + " AND ".join(where)) if where else ""
    con = db.connect()
    try:
        total = con.execute(f"SELECT count(*) FROM reconciliacion {wsql}", params).fetchone()[0]
        rows = _rows(con, f"""
            SELECT guia, cliente_real, tipo, estatus, es_retorno,
                   round(costo,2) costo, sale_price, round(sobrepeso_cobrado,2) sobrepeso,
                   round(ingreso,2) ingreso, round(margen,2) margen, mes_envio
            FROM reconciliacion {wsql}
            ORDER BY costo DESC LIMIT ? OFFSET ?""", params + [limit, offset])
        return {"total": total, "rows": rows}
    finally:
        con.close()


@app.get("/api/por-cobrar")
def por_cobrar(limit: int = 200, desde: str = "", hasta: str = ""):
    fc, fp = _fecha(desde, hasta)
    extra = "".join(" AND " + c for c in fc)
    con = db.connect()
    try:
        return _rows(con, f"""
            SELECT cliente_real, estatus, count(*) guias,
                   round(sum(costo),2) costo,
                   round(sum(costo)-sum(coalesce(ingreso,0)),2) falta_cobrar
            FROM reconciliacion
            WHERE estatus IN ('Falta cobrar (credito)','Sobrepeso pendiente','Cobrado bajo costo'){extra}
            GROUP BY cliente_real, estatus ORDER BY falta_cobrar DESC LIMIT ?""", fp + [limit])
    finally:
        con.close()


@app.get("/api/cierre/meses")
def cierre_meses():
    con = db.connect()
    try:
        meses = _rows(con, """
            SELECT mes_envio AS mes, count(*) guias,
                   round(sum(costo),2) costo,
                   round(sum(CASE WHEN costo IS NOT NULL THEN ingreso ELSE 0 END),2) ingreso,
                   round(sum(CASE WHEN costo IS NOT NULL AND ingreso IS NOT NULL THEN ingreso - costo ELSE 0 END),2) utilidad,
                   round(sum(CASE WHEN estatus IN ('Falta cobrar (credito)','Sobrepeso pendiente')
                          THEN costo - coalesce(ingreso,0) ELSE 0 END),2) por_cobrar,
                   round(sum(CASE WHEN costo IS NULL THEN coalesce(ingreso,0) ELSE 0 END),2) por_devengar
            FROM reconciliacion WHERE mes_envio IS NOT NULL AND mes_envio <> 'SIN'
            GROUP BY mes_envio ORDER BY mes_envio DESC""")
        cerrados = {r["mes"]: r for r in _rows(con, "SELECT mes, estatus, cerrado_en FROM periodos")}
        for m in meses:
            c = cerrados.get(m["mes"])
            m["estatus"] = c["estatus"] if c else "abierto"
            m["cerrado_en"] = c["cerrado_en"] if c else None
        return meses
    finally:
        con.close()


@app.get("/api/cierre/detalle")
def cierre_detalle(mes: str):
    """Detalle de un mes: cuánto se cobró por cliente y por paquetería."""
    con = db.connect()
    try:
        return _rows(con, """
            SELECT cliente_real, carrier, count(*) guias,
                   round(sum(costo),2) costo,
                   round(sum(coalesce(ingreso,0)),2) cobro,
                   round(sum(coalesce(ingreso,0)) - sum(coalesce(costo,0)),2) margen
            FROM reconciliacion WHERE mes_envio = ?
            GROUP BY cliente_real, carrier ORDER BY costo DESC NULLS LAST LIMIT 3000""", [mes])
    finally:
        con.close()


@app.get("/api/cierre/detalle/excel")
def cierre_detalle_excel(mes: str):
    """Descarga el detalle del mes (cliente × paquetería) en Excel."""
    from openpyxl import Workbook
    con = db.connect()
    try:
        rows = con.execute("""
            SELECT cliente_real, carrier, count(*) guias,
                   round(sum(costo),2), round(sum(coalesce(ingreso,0)),2),
                   round(sum(coalesce(ingreso,0)) - sum(coalesce(costo,0)),2)
            FROM reconciliacion WHERE mes_envio = ?
            GROUP BY cliente_real, carrier ORDER BY sum(costo) DESC NULLS LAST""", [mes]).fetchall()
        wb = Workbook(); ws = wb.active; ws.title = "Detalle"
        ws.append([f"Detalle de cierre · {mes}"])
        ws.append(["Cliente", "Paquetería", "Guías", "Costo", "Cobro", "Margen"])
        for r in rows:
            ws.append(list(r))
        config.EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
        out = config.EXPORTS_DIR / f"detalle_cierre_{mes}.xlsx"
        wb.save(out)
        return FileResponse(str(out), filename=out.name)
    finally:
        con.close()


@app.post("/api/cierre/cerrar")
def cerrar_mes(mes: str = Query(...)):
    con = db.connect()
    try:
        row = con.execute("""
            SELECT round(sum(costo),2), round(sum(ingreso),2),
                   round(sum(CASE WHEN ingreso IS NOT NULL THEN ingreso - costo ELSE 0 END),2)
            FROM reconciliacion WHERE mes_envio = ?""", [mes]).fetchone()
        con.execute("DELETE FROM periodos WHERE mes = ?", [mes])
        con.execute("INSERT INTO periodos VALUES (?,?,?,?,?, now())",
                    [mes, "cerrado", row[0], row[1], row[2]])
        return {"mes": mes, "estatus": "cerrado", "costo": row[0], "ingreso": row[1], "utilidad": row[2]}
    finally:
        con.close()


@app.post("/api/cierre/reabrir")
def reabrir_mes(mes: str = Query(...)):
    con = db.connect()
    try:
        con.execute("DELETE FROM periodos WHERE mes = ?", [mes])
        return {"mes": mes, "estatus": "abierto"}
    finally:
        con.close()


@app.get("/api/config-credito")
def get_config_credito():
    con = db.connect()
    try:
        return _rows(con, """
            SELECT r.seller_id, any_value(r.cliente_real) cliente, count(*) guias,
                   round(sum(r.costo),2) costo,
                   any_value(cc.metodo) metodo, any_value(cc.valor) valor, any_value(cc.nota) nota
            FROM reconciliacion r
            LEFT JOIN config_credito cc ON cc.seller_id = r.seller_id
            WHERE r.tipo = 'credito'
            GROUP BY r.seller_id ORDER BY costo DESC""")
    finally:
        con.close()


@app.post("/api/config-credito")
def set_config_credito(payload: dict):
    con = db.connect()
    try:
        sid = int(payload["seller_id"])
        con.execute("DELETE FROM config_credito WHERE seller_id = ?", [sid])
        con.execute("INSERT INTO config_credito VALUES (?,?,?,?,?)", [
            sid, payload.get("cliente"), payload.get("metodo"),
            float(payload["valor"]) if payload.get("valor") not in (None, "") else None,
            payload.get("nota")])
        return {"ok": True}
    finally:
        con.close()


@app.get("/api/tarifas/clientes")
def tarifas_clientes():
    """Clientes de crédito + su tipo de cobro (automatica/manual) + filas de tarifa."""
    con = db.connect()
    try:
        return _rows(con, """
            SELECT r.seller_id, any_value(r.cliente_real) cliente, count(*) guias,
                   round(sum(r.costo),2) costo,
                   coalesce(any_value(cc.metodo),'automatica') cobro_tipo,
                   any_value(cc.margen) margen, any_value(cc.dias_credito) dias_credito,
                   (SELECT count(*) FROM tarifas t WHERE t.seller_id = r.seller_id) filas_tarifa
            FROM reconciliacion r
            LEFT JOIN config_credito cc ON cc.seller_id = r.seller_id
            WHERE r.tipo = 'credito'
            GROUP BY r.seller_id ORDER BY costo DESC""")
    finally:
        con.close()


@app.post("/api/cliente/cobro")
def set_cobro(payload: dict):
    """Configura el cobro de un cliente de crédito: método, margen y días de crédito.
    Hace merge: solo cambia los campos enviados, conserva los demás."""
    con = db.connect()
    try:
        sid = int(payload["seller_id"])
        cur = con.execute("SELECT metodo, margen, dias_credito, cliente FROM config_credito WHERE seller_id = ?",
                          [sid]).fetchone()
        metodo = payload.get("cobro_tipo") or (cur[0] if cur else "automatica")
        margen = _f(payload["margen"]) if "margen" in payload else (cur[1] if cur else None)
        dias = _i(payload["dias_credito"]) if "dias_credito" in payload else (cur[2] if cur else None)
        cliente = payload.get("cliente") or (cur[3] if cur else None)
        con.execute("DELETE FROM config_credito WHERE seller_id = ?", [sid])
        con.execute("""INSERT INTO config_credito
            (seller_id, cliente, metodo, valor, nota, margen, dias_credito) VALUES (?,?,?,?,?,?,?)""",
            [sid, cliente, metodo, None, None, margen, dias])
        return {"ok": True}
    finally:
        con.close()


@app.get("/api/tarifas")
def get_tarifas(seller_id: int, carrier: str = "dhl"):
    con = db.connect()
    try:
        return _rows(con, """
            SELECT zona, peso_min, peso_max, precio, vigencia_desde, vigencia_hasta
            FROM tarifas WHERE seller_id = ? AND carrier = ? ORDER BY zona, peso_min""",
            [seller_id, carrier])
    finally:
        con.close()


@app.get("/api/cobranza")
def cobranza():
    """Cobranza: por cliente, lo que falta cobrar (worklist de finanzas)."""
    con = db.connect()
    try:
        return _rows(con, """
            SELECT cliente_real, any_value(tipo) tipo, count(*) guias,
                   round(sum(costo),2) costo,
                   round(sum(coalesce(ingreso,0)),2) cobrado,
                   round(sum(costo) - sum(coalesce(ingreso,0)),2) falta_cobrar,
                   round(100*sum(coalesce(ingreso,0))/nullif(sum(costo),0),1) pct_cobrado
            FROM reconciliacion
            WHERE tipo IN ('prepago','credito')
            GROUP BY cliente_real
            HAVING falta_cobrar > 1
            ORDER BY falta_cobrar DESC LIMIT 500""")
    finally:
        con.close()


@app.post("/api/tarifas")
def save_tarifas(payload: dict):
    """Reemplaza todas las filas de tarifa de un cliente (carrier dhl)."""
    sid = int(payload["seller_id"])
    cliente = payload.get("cliente")
    carrier = payload.get("carrier", "dhl")
    rows = payload.get("rows", [])
    con = db.connect()
    try:
        con.execute("DELETE FROM tarifas WHERE seller_id = ? AND carrier = ?", [sid, carrier])
        for r in rows:
            con.execute("INSERT INTO tarifas VALUES (?,?,?,?,?,?,?,?,?)", [
                sid, cliente, carrier, str(r.get("zona", "")).strip(),
                _f(r.get("peso_min")), _f(r.get("peso_max")), _f(r.get("precio")),
                r.get("vigencia_desde") or None, r.get("vigencia_hasta") or None])
        return {"ok": True, "filas": len(rows)}
    finally:
        con.close()


def _f(v):
    try:
        return float(v) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


def _i(v):
    try:
        return int(v) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


_METODOS = {"automatica", "margen_global", "margen_zona", "margen_kilo", "flat"}


# ---- Módulo Costos: rate card de la paquetería ----
@app.get("/api/costos")
def get_costos(carrier: str = "dhl"):
    con = db.connect()
    try:
        return _rows(con, """SELECT zona, peso_min, peso_max, costo, vigencia_desde, vigencia_hasta
                             FROM costos_tarifa WHERE carrier=? ORDER BY zona, peso_min""", [carrier])
    finally:
        con.close()


@app.post("/api/costos")
def save_costos(payload: dict):
    """Reemplaza el rate card de una paquetería (vigencia global para toda la matriz)."""
    carrier = payload.get("carrier", "dhl")
    vd = payload.get("vigencia_desde") or None
    vh = payload.get("vigencia_hasta") or None
    rows = payload.get("rows", [])
    con = db.connect()
    try:
        con.execute("DELETE FROM costos_tarifa WHERE carrier=?", [carrier])
        for r in rows:
            con.execute("INSERT INTO costos_tarifa VALUES (?,?,?,?,?,?,?)",
                        [carrier, str(r.get("zona", "")).strip(), _f(r.get("peso_min")),
                         _f(r.get("peso_max")), _f(r.get("costo")), vd, vh])
        return {"ok": True, "filas": len(rows)}
    finally:
        con.close()


# ---- Combustible (fuel) por periodo ----
@app.get("/api/combustible")
def get_combustible(carrier: str = "dhl"):
    con = db.connect()
    try:
        return _rows(con, """SELECT vigencia_desde, vigencia_hasta, pct FROM combustible
                             WHERE carrier=? ORDER BY vigencia_desde""", [carrier])
    finally:
        con.close()


@app.post("/api/combustible")
def save_combustible(payload: dict):
    carrier = payload.get("carrier", "dhl")
    rows = payload.get("rows", [])
    con = db.connect()
    try:
        con.execute("DELETE FROM combustible WHERE carrier=?", [carrier])
        for r in rows:
            con.execute("INSERT INTO combustible VALUES (?,?,?,?)",
                        [carrier, r.get("vigencia_desde") or None, r.get("vigencia_hasta") or None,
                         _f(r.get("pct"))])
        return {"ok": True, "filas": len(rows)}
    finally:
        con.close()


# ---- Margen por zona / por kilo (métodos de cotización) ----
@app.get("/api/margen-zona")
def get_margen_zona(seller_id: int):
    con = db.connect()
    try:
        return _rows(con, "SELECT zona, margen FROM margen_zona WHERE seller_id=? ORDER BY zona", [seller_id])
    finally:
        con.close()


@app.post("/api/margen-zona")
def save_margen_zona(payload: dict):
    sid = int(payload["seller_id"]); rows = payload.get("rows", [])
    con = db.connect()
    try:
        con.execute("DELETE FROM margen_zona WHERE seller_id=?", [sid])
        for r in rows:
            con.execute("INSERT INTO margen_zona VALUES (?,?,?)",
                        [sid, str(r.get("zona", "")).strip(), _f(r.get("margen"))])
        return {"ok": True, "filas": len(rows)}
    finally:
        con.close()


@app.get("/api/margen-kilo")
def get_margen_kilo(seller_id: int):
    con = db.connect()
    try:
        return _rows(con, "SELECT peso_min, peso_max, margen FROM margen_kilo WHERE seller_id=? ORDER BY peso_min",
                     [seller_id])
    finally:
        con.close()


@app.post("/api/margen-kilo")
def save_margen_kilo(payload: dict):
    sid = int(payload["seller_id"]); rows = payload.get("rows", [])
    con = db.connect()
    try:
        con.execute("DELETE FROM margen_kilo WHERE seller_id=?", [sid])
        for r in rows:
            con.execute("INSERT INTO margen_kilo VALUES (?,?,?)",
                        [sid, _f(r.get("peso_min")), _f(r.get("peso_max")), _f(r.get("margen"))])
        return {"ok": True, "filas": len(rows)}
    finally:
        con.close()


# ---- Preview: tarifa resultante del cliente (costo × fuel × margen × IVA) ----
@app.get("/api/tarifa-preview")
def tarifa_preview(seller_id: int, carrier: str = "dhl"):
    con = db.connect()
    try:
        cfg = con.execute("SELECT metodo, margen FROM config_credito WHERE seller_id=?", [seller_id]).fetchone()
        metodo = (cfg[0] if cfg and cfg[0] else "automatica")
        if metodo == "manual":
            metodo = "flat"
        if metodo not in _METODOS or metodo == "automatica":
            return {"metodo": metodo, "iva": config.IVA_DEFAULT, "rows": []}
        margen = cfg[1] if cfg else None
        if metodo == "flat":
            rows = _rows(con, f"""SELECT zona, peso_min, peso_max, NULL AS costo, 0 AS fuel,
                          round(precio * (1+{config.IVA_DEFAULT}),2) AS precio
                          FROM tarifas WHERE seller_id=? AND carrier=? ORDER BY zona, peso_min""",
                         [seller_id, carrier])
            return {"metodo": metodo, "iva": config.IVA_DEFAULT, "rows": rows}
        # OJO: alias externo 'cc' (distinto de 'ct' que usa el motor internamente, si no se rompe la correlación)
        expr = pricing.precio_sql(
            carrier="cc.carrier", zona="cc.zona", kilo="COALESCE(cc.peso_min,0)",
            fecha="current_date", seller=str(int(seller_id)),
            metodo="'" + metodo + "'", margen=("NULL" if margen is None else str(float(margen))))
        fuel = ("COALESCE((SELECT cb.pct FROM combustible cb WHERE cb.carrier=cc.carrier AND current_date "
                "BETWEEN COALESCE(cb.vigencia_desde,DATE '1900-01-01') AND COALESCE(cb.vigencia_hasta,DATE '2999-01-01') "
                "ORDER BY cb.vigencia_desde DESC LIMIT 1),0)")
        rows = _rows(con, f"""
            SELECT cc.zona, cc.peso_min, cc.peso_max, cc.costo,
                   round({fuel},4) AS fuel, round({expr},2) AS precio
            FROM costos_tarifa cc
            WHERE cc.carrier=? AND current_date BETWEEN COALESCE(cc.vigencia_desde,DATE '1900-01-01')
                  AND COALESCE(cc.vigencia_hasta,DATE '2999-01-01')
            ORDER BY cc.zona, cc.peso_min""", [carrier])
        return {"metodo": metodo, "iva": config.IVA_DEFAULT, "rows": rows}
    finally:
        con.close()


@app.post("/api/cobro/generar")
def generar_cobro(mes: str = Query(...)):
    """Genera el cobro por cliente del mes:
       - crédito  -> factura completa (Σ ingreso: tarifa si es manual, sistema+extra si automático)
       - prepago  -> solo extras (Σ extra), aparece únicamente si hay extras
       Upsert: conserva el ciclo (enviada/pagos) de cobros ya existentes."""
    con = db.connect()
    try:
        rows = _rows(con, """
            SELECT seller_id, any_value(cliente_real) cliente, 'credito' tipo, 'factura' concepto,
                   count(*) guias, round(sum(ingreso),2) monto
            FROM reconciliacion
            WHERE mes_envio = ? AND seller_id IS NOT NULL AND tipo='credito' AND ingreso IS NOT NULL
            GROUP BY seller_id HAVING sum(ingreso) > 0
            UNION ALL
            SELECT seller_id, any_value(cliente_real), 'prepago', 'extra',
                   count(*) FILTER (WHERE extra > 0.5), round(sum(extra),2)
            FROM reconciliacion
            WHERE mes_envio = ? AND seller_id IS NOT NULL AND tipo='prepago'
            GROUP BY seller_id HAVING sum(extra) > 0.5
            ORDER BY monto DESC""", [mes, mes])
        keep = []
        for r in rows:
            keep.append(r["seller_id"])
            exists = con.execute("SELECT 1 FROM cobros WHERE seller_id=? AND mes=?",
                                 [r["seller_id"], mes]).fetchone()
            if exists:
                con.execute("""UPDATE cobros SET cliente=?, guias=?, monto=?, tipo=?, concepto=?
                               WHERE seller_id=? AND mes=?""",
                            [r["cliente"], r["guias"], r["monto"], r["tipo"], r["concepto"], r["seller_id"], mes])
            else:
                con.execute("""INSERT INTO cobros
                    (seller_id,cliente,mes,guias,monto,estatus,generado_en,nota,tipo,concepto,monto_pagado,fecha_enviada,fecha_vencimiento)
                    VALUES (?,?,?,?,?,'generado', now(), NULL, ?,?, 0, NULL, NULL)""",
                    [r["seller_id"], r["cliente"], mes, r["guias"], r["monto"], r["tipo"], r["concepto"]])
        # quita cobros obsoletos del mes (clientes que ya no aplican) que sigan sin tocar
        if keep:
            ph = ",".join("?" * len(keep))
            con.execute(f"DELETE FROM cobros WHERE mes=? AND estatus='generado' AND seller_id NOT IN ({ph})",
                        [mes, *keep])
        else:
            con.execute("DELETE FROM cobros WHERE mes=? AND estatus='generado'", [mes])
        tot = con.execute("SELECT count(*), round(sum(monto),2) FROM cobros WHERE mes=?", [mes]).fetchone()
        return {"mes": mes, "sellers": tot[0], "monto_total": tot[1] or 0}
    finally:
        con.close()


@app.get("/api/cobros")
def list_cobros(mes: str):
    """Libro de cuentas por cobrar del mes: monto, pagado, saldo, vencimiento y atraso."""
    con = db.connect()
    try:
        return _rows(con, """
            WITH pg AS (SELECT seller_id, round(sum(monto),2) pagado FROM pagos WHERE mes=? GROUP BY seller_id)
            SELECT c.seller_id, c.cliente, c.tipo, c.concepto, c.guias, c.monto,
                   COALESCE(pg.pagado, 0) AS pagado,
                   round(c.monto - COALESCE(pg.pagado, 0), 2) AS saldo,
                   c.estatus, c.fecha_enviada, c.fecha_vencimiento,
                   CASE WHEN c.fecha_vencimiento IS NOT NULL AND c.estatus <> 'pagado'
                             AND current_date > c.fecha_vencimiento
                        THEN date_diff('day', c.fecha_vencimiento, current_date) ELSE 0 END AS dias_atraso,
                   (SELECT count(*) FROM cobro_adjuntos a WHERE a.seller_id=c.seller_id AND a.mes=c.mes) > 0 AS tiene_pdf
            FROM cobros c LEFT JOIN pg ON pg.seller_id = c.seller_id
            WHERE c.mes = ? ORDER BY saldo DESC, c.monto DESC""", [mes, mes])
    finally:
        con.close()


@app.post("/api/cobro/enviar")
def cobro_enviar(payload: dict):
    """Marca la factura como enviada y calcula vencimiento = fecha + días de crédito del cliente."""
    sid = int(payload["seller_id"]); mes = payload["mes"]
    con = db.connect()
    try:
        row = con.execute("SELECT dias_credito FROM config_credito WHERE seller_id=?", [sid]).fetchone()
        dias = (row[0] if row and row[0] else config.DIAS_CREDITO_DEFAULT)
        fenv = dt.date.fromisoformat(payload["fecha"]) if payload.get("fecha") else dt.date.today()
        fvenc = fenv + dt.timedelta(days=int(dias))
        con.execute("""UPDATE cobros SET fecha_enviada=?, fecha_vencimiento=?,
                       estatus = CASE WHEN estatus='pagado' THEN 'pagado'
                                      WHEN COALESCE(monto_pagado,0) > 0 THEN 'parcial' ELSE 'enviado' END
                       WHERE seller_id=? AND mes=?""", [fenv, fvenc, sid, mes])
        return {"ok": True, "dias_credito": int(dias), "vencimiento": str(fvenc)}
    finally:
        con.close()


def _recompute_cobro(con, sid, mes) -> float:
    """Recalcula monto_pagado y estatus de un cobro a partir de sus pagos.
    Si no hay pagos, conserva el estatus del ciclo (generado/enviado/aprobado)."""
    pagado = con.execute("SELECT COALESCE(round(sum(monto),2),0) FROM pagos WHERE seller_id=? AND mes=?",
                         [sid, mes]).fetchone()[0]
    row = con.execute("SELECT monto, estatus FROM cobros WHERE seller_id=? AND mes=?", [sid, mes]).fetchone()
    if not row:
        return pagado
    total, cur = (row[0] or 0), row[1]
    if pagado >= total - 0.5:
        est = "pagado"
    elif pagado > 0:
        est = "parcial"
    else:
        est = cur if cur in ("generado", "enviado", "aprobado") else "enviado"
    con.execute("UPDATE cobros SET monto_pagado=?, estatus=? WHERE seller_id=? AND mes=?", [pagado, est, sid, mes])
    return pagado


@app.post("/api/cobro/pago")
def cobro_pago(payload: dict):
    """Registra un abono (pago parcial o total). Recalcula saldo y estatus."""
    sid = int(payload["seller_id"]); mes = payload["mes"]
    monto = float(payload["monto"])
    fecha = dt.date.fromisoformat(payload["fecha"]) if payload.get("fecha") else dt.date.today()
    con = db.connect()
    try:
        con.execute("INSERT INTO pagos VALUES (?,?,?,?,?, now())",
                    [sid, mes, fecha, monto, payload.get("nota")])
        pagado = _recompute_cobro(con, sid, mes)
        trow = con.execute("SELECT monto FROM cobros WHERE seller_id=? AND mes=?", [sid, mes]).fetchone()
        total = (trow[0] if trow else 0) or 0
        return {"ok": True, "pagado": pagado, "saldo": round(total - pagado, 2)}
    finally:
        con.close()


@app.get("/api/cobro/pagos")
def list_pagos(seller_id: int, mes: str):
    """Histórico de abonos de un cliente/mes."""
    con = db.connect()
    try:
        return _rows(con, """SELECT rowid AS id, CAST(fecha AS VARCHAR) fecha, monto, nota,
                                    CAST(registrado_en AS VARCHAR) registrado_en
                             FROM pagos WHERE seller_id=? AND mes=? ORDER BY fecha, registrado_en""",
                     [seller_id, mes])
    finally:
        con.close()


@app.post("/api/cobro/pago/eliminar")
def del_pago(payload: dict):
    """Elimina un abono (corrección) y recalcula el cobro."""
    sid = int(payload["seller_id"]); mes = payload["mes"]; pid = int(payload["id"])
    con = db.connect()
    try:
        con.execute("DELETE FROM pagos WHERE rowid=? AND seller_id=? AND mes=?", [pid, sid, mes])
        _recompute_cobro(con, sid, mes)
        return {"ok": True}
    finally:
        con.close()


@app.post("/api/cobro/estatus")
def cobro_estatus(payload: dict):
    con = db.connect()
    try:
        con.execute("UPDATE cobros SET estatus = ? WHERE seller_id = ? AND mes = ?",
                    [payload["estatus"], int(payload["seller_id"]), payload["mes"]])
        return {"ok": True}
    finally:
        con.close()


@app.get("/api/cobro/detalle")
def cobro_detalle(seller_id: int, mes: str):
    """Detalle del cobro de un cliente/mes: resumen + guías cobradas (para 'Ver detalle')."""
    con = db.connect()
    try:
        c = con.execute("""SELECT cliente, tipo, concepto, monto, estatus, fecha_enviada, fecha_vencimiento
                           FROM cobros WHERE seller_id=? AND mes=?""", [seller_id, mes]).fetchone()
        if not c:
            return JSONResponse({"error": "sin cobro"}, 404)
        pagado = con.execute("SELECT COALESCE(round(sum(monto),2),0) FROM pagos WHERE seller_id=? AND mes=?",
                             [seller_id, mes]).fetchone()[0]
        es_extra = c[2] == "extra"
        imp = "r.extra" if es_extra else "r.ingreso"
        filtro = "AND r.extra > 0.5" if es_extra else ""
        guias = _rows(con, f"""
            WITH carrier AS (SELECT * FROM (SELECT *, row_number() OVER (PARTITION BY guia ORDER BY fecha_factura NULLS LAST) rn
                             FROM facturas_carrier) WHERE rn=1)
            SELECT fc.guia, CAST(fc.fecha_envio AS VARCHAR) fecha, fc.producto, fc.destino, fc.kilos, fc.zona,
                   r.es_retorno, round({imp},2) importe
            FROM reconciliacion r JOIN carrier fc ON fc.guia = r.guia
            WHERE r.seller_id=? AND r.mes_envio=? {filtro} ORDER BY importe DESC LIMIT 500""", [seller_id, mes])
        return {"cliente": c[0], "tipo": c[1], "concepto": c[2], "monto": c[3], "estatus": c[4],
                "fecha_enviada": str(c[5]) if c[5] else None,
                "fecha_vencimiento": str(c[6]) if c[6] else None,
                "pagado": pagado, "saldo": round((c[3] or 0) - pagado, 2), "guias": guias}
    finally:
        con.close()


_CARRIER_LABEL = {"dhl": "DHL", "fedex": "FedEx", "paquete_express": "Paquete Express"}
_COBRO_HEADERS = ["No.De Guia", "Referencia", "Producto", "Org", "Des", "Zona", "Pza", "Kilos", "kilo",
                  "Servicio", "Fecha Envio", "Flete", "Otros costos", "Subtotal", "IVA", "Total",
                  "Seguro", "% IVA", "Moneda", "Tipo de Cambio", "Remitente", "Destinatario"]


@app.get("/api/cobro/seller")
def cobro_seller(seller_id: int, mes: str):
    """Excel de cobro al cliente (formato nuevo con desglose), UNA HOJA POR PAQUETERÍA.
    Sin costo ni margen: solo el servicio + el precio a cobrar (Flete/Subtotal/IVA/Total)."""
    from collections import OrderedDict
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    con = db.connect()
    try:
        cliente = con.execute("SELECT any_value(cliente_real) FROM reconciliacion WHERE seller_id = ?",
                              [seller_id]).fetchone()[0] or str(seller_id)
        crow = con.execute("SELECT concepto FROM cobros WHERE seller_id=? AND mes=?", [seller_id, mes]).fetchone()
        es_extra = bool(crow and crow[0] == "extra")          # prepago: solo guías con extra
        imp = "r.extra" if es_extra else "r.ingreso"
        filtro = "AND r.extra > 0.5" if es_extra else ""
        rows = con.execute(f"""
            WITH carrier AS (SELECT * FROM (SELECT *, row_number() OVER (PARTITION BY guia ORDER BY fecha_factura NULLS LAST) rn
                             FROM facturas_carrier) WHERE rn = 1)
            SELECT fc.carrier, fc.guia, fc.referencia, fc.producto, fc.origen, fc.destino, fc.zona,
                   fc.piezas, fc.kilos, CAST(ceil(fc.kilos) AS INTEGER) AS kilo, fc.fecha_envio,
                   fc.remitente, fc.destinatario, round({imp},2) AS total
            FROM reconciliacion r JOIN carrier fc ON fc.guia = r.guia
            WHERE r.seller_id = ? AND r.mes_envio = ? {filtro} ORDER BY fc.carrier, fc.guia""",
            [seller_id, mes]).fetchall()
        iva_r = config.IVA_DEFAULT
        bycar = OrderedDict()
        for r in rows:
            bycar.setdefault(r[0], []).append(r)
        wb = Workbook(); wb.remove(wb.active)
        if not bycar:
            ws = wb.create_sheet("Cobro"); ws.append([f"Cobro a {cliente} · {mes}"]); ws.append(_COBRO_HEADERS)
        for car, crows in bycar.items():
            label = _CARRIER_LABEL.get(car, str(car).upper())
            ws = wb.create_sheet(label[:31])
            ws.append([f"Cobro a {cliente} · {mes} · {label}"])
            ws.append(_COBRO_HEADERS)
            tot = 0.0
            for r in crows:
                total = r[13] or 0.0
                sub = round(total / (1 + iva_r), 2)
                ivamt = round(total - sub, 2)
                tot += total
                ws.append([r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8], r[9], r[3], r[10],
                           sub, 0, sub, ivamt, round(total, 2), 0, round(iva_r * 100), "MXN", 1, r[11], r[12]])
            ws.append(["TOTAL"] + [""] * 14 + [round(tot, 2)] + [""] * 6)
            for ci in (12, 14, 15, 16, 17):  # Flete, Subtotal, IVA, Total, Seguro
                for cell in ws[get_column_letter(ci)][2:]:
                    cell.number_format = '#,##0.00'
        config.EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
        safe = "".join(ch for ch in str(cliente) if ch.isalnum() or ch in " -_")[:40].strip() or str(seller_id)
        out = config.EXPORTS_DIR / f"cobro_{safe}_{mes}.xlsx"
        wb.save(out)
        return FileResponse(str(out), filename=out.name)
    finally:
        con.close()


@app.post("/api/cobro/pdf")
def subir_pdf(seller_id: int = Query(...), mes: str = Query(...), file: UploadFile = File(...)):
    """Adjunta el PDF de la factura de un cliente (Supabase Storage, o disco local)."""
    content = file.file.read()
    if supa.enabled():
        path = f"{seller_id}/{mes}/{file.filename}"
        supa.upload(path, content, "application/pdf")
    else:
        dest_dir = config.DATA_DIR / "uploads" / "pdf"
        dest_dir.mkdir(parents=True, exist_ok=True)
        path = str(dest_dir / f"{seller_id}_{mes}_{file.filename}")
        with open(path, "wb") as f:
            f.write(content)
    con = db.connect()
    try:
        con.execute("DELETE FROM cobro_adjuntos WHERE seller_id = ? AND mes = ?", [seller_id, mes])
        con.execute("INSERT INTO cobro_adjuntos VALUES (?,?,?,?, now())",
                    [seller_id, mes, file.filename, path])
        return {"ok": True, "filename": file.filename}
    finally:
        con.close()


@app.get("/api/cobro/pdf")
def ver_pdf(seller_id: int, mes: str):
    con = db.connect()
    try:
        row = con.execute("SELECT path, filename FROM cobro_adjuntos WHERE seller_id = ? AND mes = ?",
                          [seller_id, mes]).fetchone()
        if not row:
            return JSONResponse({"error": "sin PDF"}, 404)
        if supa.enabled():
            url = supa.signed_url(row[0])
            return RedirectResponse(url) if url else JSONResponse({"error": "no se pudo firmar"}, 500)
        return FileResponse(row[0], filename=row[1], media_type="application/pdf")
    finally:
        con.close()


@app.get("/api/export")
def download_export():
    con = db.connect()
    try:
        out = export.export(con)
        return FileResponse(out, filename="conciliacion.xlsx")
    finally:
        con.close()


# ---------- UI estática ----------
app.mount("/", StaticFiles(directory=str(UI_DIR), html=True), name="ui")
